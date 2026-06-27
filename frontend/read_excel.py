import subprocess
import sys
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])

import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\New Microsoft Excel Worksheet (3).xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Merged cells: {ws.merged_cells.ranges}")
    print()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                row_data.append(f"[{cell.coordinate}] {val}")
        if row_data:
            print(" | ".join(row_data))
